"""Results processing and analysis for ML Agents experiments.

This module provides comprehensive results processing, including database
persistence, statistical analysis, and export functionality for multiple formats.
"""

import json
import logging
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd

from ..core.database_manager import DatabaseConfig, DatabaseManager
from ..reasoning.base import StandardResponse

logger = logging.getLogger(__name__)


@dataclass
class ExperimentSummary:
    """Summary statistics for an experiment."""

    experiment_id: str
    total_runs: int
    completed_runs: int
    failed_runs: int
    accuracy: float
    avg_execution_time_ms: float
    total_cost: float
    approaches_tested: List[str]
    models_used: List[str]
    parsing_success_rate: float
    created_at: datetime
    last_updated: datetime


@dataclass
class ApproachComparison:
    """Comparison metrics between reasoning approaches."""

    approach_name: str
    accuracy: float
    avg_execution_time_ms: float
    total_cost: float
    sample_count: int
    parsing_success_rate: float
    confidence_avg: float
    failure_rate: float


class ResultsProcessor:
    """Processes and analyzes experiment results with database persistence."""

    def __init__(self, db_config: Optional[DatabaseConfig] = None):
        """Initialize results processor.

        Args:
            db_config: Database configuration (uses defaults if None)
        """
        self.db_config = db_config or DatabaseConfig()
        self.db_manager = DatabaseManager(self.db_config)

    # ==================== Persistence Methods ====================

    def save_experiment(
        self,
        experiment_id: str,
        name: str,
        config: Dict[str, Any],
        description: Optional[str] = None,
    ) -> None:
        """Save experiment metadata to database.

        Args:
            experiment_id: Unique experiment identifier
            name: Experiment name
            config: Experiment configuration
            description: Optional experiment description
        """
        with self.db_manager.get_connection() as conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO experiments (id, name, description, config_json, status)
                VALUES (?, ?, ?, ?, 'running')
            """,
                (experiment_id, name, description, json.dumps(config)),
            )
            conn.commit()

        logger.info(f"Saved experiment {experiment_id}: {name}")

    def _calculate_correctness(
        self, extracted_answer: str, expected_answer: str
    ) -> bool:
        """Calculate if the extracted answer is correct.

        Args:
            extracted_answer: The answer extracted from model output
            expected_answer: The expected/ground truth answer

        Returns:
            True if answers match (case-insensitive, stripped)
        """
        if not extracted_answer or not expected_answer:
            return False

        # Normalize both answers for comparison
        extracted = str(extracted_answer).strip().lower()
        expected = str(expected_answer).strip().lower()

        return extracted == expected

    def save_run_result(self, result: StandardResponse) -> None:
        """Save a single run result to database.

        Args:
            result: StandardResponse from reasoning execution
        """
        # Generate run ID if not present
        run_id = result.metadata.get("run_id", str(uuid.uuid4()))

        # Extract data from StandardResponse
        # Calculate correctness by comparing extracted answer with expected answer
        expected_answer = result.metadata.get("expected_answer", "")
        extracted_answer = result.extracted_answer or ""
        is_correct = self._calculate_correctness(extracted_answer, expected_answer)

        run_data = {
            "id": run_id,
            "experiment_id": result.metadata.get("experiment_id", "unknown"),
            "approach_name": result.metadata.get("approach", "unknown"),
            "provider": result.metadata.get("provider", "unknown"),
            "model": result.metadata.get("model", "unknown"),
            "sample_index": result.metadata.get("sample_index", 0),
            "input_text": result.metadata.get("input_text", ""),
            "expected_answer": expected_answer,
            "raw_output": result.text,  # Changed from result.response to result.text
            "parsed_answer": extracted_answer,  # Changed from result.parsed_answer to result.extracted_answer
            "parsing_method": result.metadata.get("parsing_method", "none"),
            "parsing_confidence": result.metadata.get("parsing_confidence", 0.0),
            "is_correct": is_correct,  # Now calculated from comparison
            "execution_time_ms": int(result.metadata.get("latency", 0) * 1000),
            "cost_estimate": result.metadata.get("cost", 0.0),
            "metadata_json": json.dumps(result.metadata),
        }

        try:
            with self.db_manager.get_connection() as conn:
                # Save run
                conn.execute(
                    """
                    INSERT OR REPLACE INTO runs (
                        id, experiment_id, approach_name, provider, model,
                        sample_index, input_text, expected_answer, raw_output,
                        parsed_answer, parsing_method, parsing_confidence,
                        is_correct, execution_time_ms, cost_estimate, metadata_json
                    ) VALUES (
                        :id, :experiment_id, :approach_name, :provider, :model,
                        :sample_index, :input_text, :expected_answer, :raw_output,
                        :parsed_answer, :parsing_method, :parsing_confidence,
                        :is_correct, :execution_time_ms, :cost_estimate, :metadata_json
                    )
                """,
                    run_data,
                )

                # Save parsing metrics if available
                if "parsing_metrics" in result.metadata:
                    self.save_parsing_metrics(
                        run_id, result.metadata["parsing_metrics"], conn
                    )

                conn.commit()
                logger.debug(f"Saved run result {run_id} to database")

        except Exception as e:
            logger.error(f"Failed to save run result {run_id} to database: {e}")
            logger.debug(f"Run data: {run_data}")
            # Don't re-raise - handle errors gracefully

        # Check for auto-backup
        self.db_manager.check_auto_backup()

    def save_parsing_metrics(
        self, run_id: str, metrics: Dict[str, Any], conn: Optional[Any] = None
    ) -> None:
        """Save parsing metrics for a run.

        Args:
            run_id: Run identifier
            metrics: Parsing metrics dictionary
            conn: Optional database connection to reuse
        """
        parsing_data = {
            "run_id": run_id,
            "parsing_attempts": metrics.get("attempts", 0),
            "fallback_used": metrics.get("fallback_used", False),
            "confidence_score": metrics.get("confidence", 0.0),
            "extraction_time_ms": metrics.get("extraction_time_ms", 0),
            "error_details": metrics.get("error_details", ""),
        }

        if conn:
            conn.execute(
                """
                INSERT OR REPLACE INTO parsing_metrics (
                    run_id, parsing_attempts, fallback_used,
                    confidence_score, extraction_time_ms, error_details
                ) VALUES (
                    :run_id, :parsing_attempts, :fallback_used,
                    :confidence_score, :extraction_time_ms, :error_details
                )
            """,
                parsing_data,
            )
        else:
            with self.db_manager.get_connection() as conn:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO parsing_metrics (
                        run_id, parsing_attempts, fallback_used,
                        confidence_score, extraction_time_ms, error_details
                    ) VALUES (
                        :run_id, :parsing_attempts, :fallback_used,
                        :confidence_score, :extraction_time_ms, :error_details
                    )
                """,
                    parsing_data,
                )
                conn.commit()

    def update_experiment_status(self, experiment_id: str, status: str) -> None:
        """Update experiment status.

        Args:
            experiment_id: Experiment identifier
            status: New status ('running', 'completed', 'failed')
        """
        with self.db_manager.get_connection() as conn:
            conn.execute(
                "UPDATE experiments SET status = ? WHERE id = ?",
                (status, experiment_id),
            )
            conn.commit()

    # ==================== Analysis Methods ====================

    def get_experiment_summary(self, experiment_id: str) -> Optional[ExperimentSummary]:
        """Get summary statistics for an experiment.

        Args:
            experiment_id: Experiment identifier

        Returns:
            ExperimentSummary or None if experiment not found
        """
        with self.db_manager.get_connection() as conn:
            # Get experiment metadata
            cursor = conn.cursor()
            cursor.execute(
                "SELECT name, created_at, status FROM experiments WHERE id = ?",
                (experiment_id,),
            )
            exp_data = cursor.fetchone()

            if not exp_data:
                return None

            # Get run statistics
            cursor.execute(
                """
                SELECT
                    COUNT(*) as total_runs,
                    SUM(CASE WHEN is_correct IS NOT NULL THEN 1 ELSE 0 END) as completed_runs,
                    SUM(CASE WHEN is_correct IS NULL THEN 1 ELSE 0 END) as failed_runs,
                    AVG(CASE WHEN is_correct THEN 1.0 ELSE 0.0 END) as accuracy,
                    AVG(execution_time_ms) as avg_execution_time,
                    SUM(cost_estimate) as total_cost,
                    COUNT(DISTINCT approach_name) as approach_count,
                    COUNT(DISTINCT model) as model_count
                FROM runs
                WHERE experiment_id = ?
            """,
                (experiment_id,),
            )

            stats = cursor.fetchone()

            # Get approach and model lists
            cursor.execute(
                "SELECT DISTINCT approach_name FROM runs WHERE experiment_id = ?",
                (experiment_id,),
            )
            approaches = [row[0] for row in cursor.fetchall()]

            cursor.execute(
                "SELECT DISTINCT model FROM runs WHERE experiment_id = ?",
                (experiment_id,),
            )
            models = [row[0] for row in cursor.fetchall()]

            # Get parsing success rate
            cursor.execute(
                """
                SELECT AVG(CASE WHEN pm.confidence_score > 0.5 THEN 1.0 ELSE 0.0 END) as parsing_success_rate
                FROM runs r
                LEFT JOIN parsing_metrics pm ON r.id = pm.run_id
                WHERE r.experiment_id = ?
            """,
                (experiment_id,),
            )

            parsing_rate = cursor.fetchone()[0] or 0.0

            # Get last update time
            cursor.execute(
                "SELECT MAX(created_at) FROM runs WHERE experiment_id = ?",
                (experiment_id,),
            )
            last_updated = cursor.fetchone()[0] or exp_data["created_at"]

            return ExperimentSummary(
                experiment_id=experiment_id,
                total_runs=stats["total_runs"] or 0,
                completed_runs=stats["completed_runs"] or 0,
                failed_runs=stats["failed_runs"] or 0,
                accuracy=stats["accuracy"] or 0.0,
                avg_execution_time_ms=stats["avg_execution_time"] or 0.0,
                total_cost=stats["total_cost"] or 0.0,
                approaches_tested=approaches,
                models_used=models,
                parsing_success_rate=parsing_rate,
                created_at=datetime.fromisoformat(exp_data["created_at"]),
                last_updated=datetime.fromisoformat(last_updated),
            )

    def compare_approaches(
        self, experiment_id: str, approaches: Optional[List[str]] = None
    ) -> List[ApproachComparison]:
        """Compare performance metrics across reasoning approaches.

        Args:
            experiment_id: Experiment identifier
            approaches: Optional list of approaches to compare (all if None)

        Returns:
            List of ApproachComparison objects
        """
        with self.db_manager.get_connection() as conn:
            query = """
                SELECT
                    r.approach_name,
                    COUNT(*) as sample_count,
                    AVG(CASE WHEN r.is_correct THEN 1.0 ELSE 0.0 END) as accuracy,
                    AVG(r.execution_time_ms) as avg_execution_time,
                    SUM(r.cost_estimate) as total_cost,
                    AVG(r.parsing_confidence) as confidence_avg,
                    AVG(CASE WHEN r.is_correct IS NULL THEN 1.0 ELSE 0.0 END) as failure_rate,
                    AVG(CASE WHEN pm.confidence_score > 0.5 THEN 1.0 ELSE 0.0 END) as parsing_success_rate
                FROM runs r
                LEFT JOIN parsing_metrics pm ON r.id = pm.run_id
                WHERE r.experiment_id = ?
            """

            params = [experiment_id]

            if approaches:
                placeholders = ",".join(["?" for _ in approaches])
                query += f" AND r.approach_name IN ({placeholders})"
                params.extend(approaches)

            query += " GROUP BY r.approach_name ORDER BY accuracy DESC"

            cursor = conn.cursor()
            cursor.execute(query, params)

            comparisons = []
            for row in cursor.fetchall():
                comparisons.append(
                    ApproachComparison(
                        approach_name=row["approach_name"],
                        accuracy=row["accuracy"] or 0.0,
                        avg_execution_time_ms=row["avg_execution_time"] or 0.0,
                        total_cost=row["total_cost"] or 0.0,
                        sample_count=row["sample_count"],
                        parsing_success_rate=row["parsing_success_rate"] or 0.0,
                        confidence_avg=row["confidence_avg"] or 0.0,
                        failure_rate=row["failure_rate"] or 0.0,
                    )
                )

            return comparisons

    def generate_accuracy_report(self, experiment_id: str) -> Dict[str, Any]:
        """Generate detailed accuracy report for an experiment.

        Args:
            experiment_id: Experiment identifier

        Returns:
            Dictionary with accuracy analysis
        """
        with self.db_manager.get_connection() as conn:
            # Overall accuracy by approach
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    approach_name,
                    COUNT(*) as total,
                    SUM(CASE WHEN is_correct THEN 1 ELSE 0 END) as correct,
                    AVG(CASE WHEN is_correct THEN 1.0 ELSE 0.0 END) as accuracy
                FROM runs
                WHERE experiment_id = ? AND is_correct IS NOT NULL
                GROUP BY approach_name
            """,
                (experiment_id,),
            )

            accuracy_by_approach = {
                row["approach_name"]: {
                    "total": row["total"],
                    "correct": row["correct"],
                    "accuracy": row["accuracy"],
                }
                for row in cursor.fetchall()
            }

            # Accuracy by model
            cursor.execute(
                """
                SELECT
                    model,
                    AVG(CASE WHEN is_correct THEN 1.0 ELSE 0.0 END) as accuracy,
                    COUNT(*) as sample_count
                FROM runs
                WHERE experiment_id = ? AND is_correct IS NOT NULL
                GROUP BY model
            """,
                (experiment_id,),
            )

            accuracy_by_model = {
                row["model"]: {
                    "accuracy": row["accuracy"],
                    "sample_count": row["sample_count"],
                }
                for row in cursor.fetchall()
            }

            # Parsing method effectiveness
            cursor.execute(
                """
                SELECT
                    parsing_method,
                    COUNT(*) as count,
                    AVG(parsing_confidence) as avg_confidence,
                    AVG(CASE WHEN is_correct THEN 1.0 ELSE 0.0 END) as accuracy
                FROM runs
                WHERE experiment_id = ? AND parsing_method IS NOT NULL
                GROUP BY parsing_method
            """,
                (experiment_id,),
            )

            parsing_effectiveness = {
                row["parsing_method"]: {
                    "count": row["count"],
                    "avg_confidence": row["avg_confidence"],
                    "accuracy": row["accuracy"],
                }
                for row in cursor.fetchall()
            }

            return {
                "experiment_id": experiment_id,
                "accuracy_by_approach": accuracy_by_approach,
                "accuracy_by_model": accuracy_by_model,
                "parsing_effectiveness": parsing_effectiveness,
                "generated_at": datetime.now().isoformat(),
            }

    def identify_failure_patterns(
        self, experiment_id: str, limit: int = 20
    ) -> List[Dict[str, Any]]:
        """Identify common failure patterns in an experiment.

        Args:
            experiment_id: Experiment identifier
            limit: Maximum number of failures to analyze

        Returns:
            List of failure patterns with examples
        """
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()

            # Get failed runs with details
            cursor.execute(
                """
                SELECT
                    r.approach_name,
                    r.model,
                    r.input_text,
                    r.expected_answer,
                    r.parsed_answer,
                    r.parsing_method,
                    pm.error_details
                FROM runs r
                LEFT JOIN parsing_metrics pm ON r.id = pm.run_id
                WHERE r.experiment_id = ? AND r.is_correct = 0
                LIMIT ?
            """,
                (experiment_id, limit),
            )

            failures = []
            for row in cursor.fetchall():
                failure = {
                    "approach": row["approach_name"],
                    "model": row["model"],
                    "input": (
                        row["input_text"][:100] + "..."
                        if len(row["input_text"]) > 100
                        else row["input_text"]
                    ),
                    "expected": row["expected_answer"],
                    "parsed": row["parsed_answer"],
                    "parsing_method": row["parsing_method"],
                    "error": row["error_details"],
                }

                # Categorize failure type
                if not row["parsed_answer"]:
                    failure["failure_type"] = "parsing_failed"
                elif row["error_details"]:
                    failure["failure_type"] = "execution_error"
                else:
                    failure["failure_type"] = "incorrect_answer"

                failures.append(failure)

            # Aggregate failure patterns
            patterns = {}
            for failure in failures:
                key = (failure["failure_type"], failure["approach"], failure["model"])
                if key not in patterns:
                    patterns[key] = {
                        "failure_type": failure["failure_type"],
                        "approach": failure["approach"],
                        "model": failure["model"],
                        "count": 0,
                        "examples": [],
                    }
                patterns[key]["count"] += 1
                if len(patterns[key]["examples"]) < 3:
                    patterns[key]["examples"].append(
                        {
                            "input": failure["input"],
                            "expected": failure["expected"],
                            "parsed": failure["parsed"],
                        }
                    )

            return list(patterns.values())

    # ==================== Export Methods ====================

    def export_to_csv(self, experiment_id: str, output_path: str) -> None:
        """Export experiment results to CSV format.

        Args:
            experiment_id: Experiment identifier
            output_path: Path for output CSV file
        """
        with self.db_manager.get_connection() as conn:
            # Query all runs with parsing metrics
            query = """
                SELECT
                    r.*,
                    pm.parsing_attempts,
                    pm.fallback_used,
                    pm.confidence_score as parsing_confidence_score,
                    pm.extraction_time_ms as parsing_time_ms
                FROM runs r
                LEFT JOIN parsing_metrics pm ON r.id = pm.run_id
                WHERE r.experiment_id = ?
                ORDER BY r.created_at
            """

            df = pd.read_sql_query(query, conn, params=(experiment_id,))

            # Clean up metadata JSON column
            if "metadata_json" in df.columns:
                df.drop("metadata_json", axis=1, inplace=True)

            # Save to CSV
            df.to_csv(output_path, index=False)
            logger.info(f"Exported {len(df)} runs to {output_path}")

    def export_to_excel(self, experiment_ids: List[str], output_path: str) -> None:
        """Export multiple experiments to Excel with formatting.

        Args:
            experiment_ids: List of experiment identifiers
            output_path: Path for output Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_data = []

        for exp_id in experiment_ids:
            summary = self.get_experiment_summary(exp_id)
            if summary:
                summary_data.append(
                    {
                        "Experiment ID": summary.experiment_id,
                        "Total Runs": summary.total_runs,
                        "Accuracy": f"{summary.accuracy:.2%}",
                        "Avg Time (ms)": f"{summary.avg_execution_time_ms:.2f}",
                        "Total Cost": f"${summary.total_cost:.4f}",
                        "Approaches": ", ".join(summary.approaches_tested),
                        "Models": ", ".join(summary.models_used),
                    }
                )

        # Write summary data
        if summary_data:
            df_summary = pd.DataFrame(summary_data)
            for r_idx, row in enumerate(
                dataframe_to_rows(df_summary, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    cell = summary_sheet.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                        )

        # Individual experiment sheets
        for exp_id in experiment_ids:
            with self.db_manager.get_connection() as conn:
                query = """
                    SELECT
                        r.approach_name,
                        r.model,
                        r.sample_index,
                        r.input_text,
                        r.expected_answer,
                        r.parsed_answer,
                        r.is_correct,
                        r.parsing_confidence,
                        r.execution_time_ms,
                        r.cost_estimate
                    FROM runs r
                    WHERE r.experiment_id = ?
                    ORDER BY r.approach_name, r.sample_index
                """

                df = pd.read_sql_query(query, conn, params=(exp_id,))

                if not df.empty:
                    sheet = wb.create_sheet(f"Exp_{exp_id[:8]}")

                    # Write data with formatting
                    for r_idx, row in enumerate(
                        dataframe_to_rows(df, index=False, header=True), 1
                    ):
                        for c_idx, value in enumerate(row, 1):
                            cell = sheet.cell(row=r_idx, column=c_idx, value=value)

                            if r_idx == 1:  # Header row
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(
                                    start_color="CCCCCC",
                                    end_color="CCCCCC",
                                    fill_type="solid",
                                )
                            elif c_idx == 7 and r_idx > 1:  # is_correct column
                                if value:
                                    cell.fill = PatternFill(
                                        start_color="90EE90",
                                        end_color="90EE90",
                                        fill_type="solid",
                                    )
                                else:
                                    cell.fill = PatternFill(
                                        start_color="FFB6C1",
                                        end_color="FFB6C1",
                                        fill_type="solid",
                                    )

                    # Auto-adjust column widths
                    for column in sheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column[0].column_letter].width = (
                            adjusted_width
                        )

        # Comparison sheet
        if len(experiment_ids) > 1:
            comparison_sheet = wb.create_sheet("Comparison")
            comparison_data = []

            for exp_id in experiment_ids:
                comparisons = self.compare_approaches(exp_id)
                for comp in comparisons:
                    comparison_data.append(
                        {
                            "Experiment": exp_id[:8],
                            "Approach": comp.approach_name,
                            "Accuracy": f"{comp.accuracy:.2%}",
                            "Avg Time (ms)": f"{comp.avg_execution_time_ms:.2f}",
                            "Total Cost": f"${comp.total_cost:.4f}",
                            "Samples": comp.sample_count,
                        }
                    )

            if comparison_data:
                df_comp = pd.DataFrame(comparison_data)
                for r_idx, row in enumerate(
                    dataframe_to_rows(df_comp, index=False, header=True), 1
                ):
                    for c_idx, value in enumerate(row, 1):
                        cell = comparison_sheet.cell(
                            row=r_idx, column=c_idx, value=value
                        )
                        if r_idx == 1:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color="CCCCCC",
                                end_color="CCCCCC",
                                fill_type="solid",
                            )

        # Save workbook
        wb.save(output_path)
        logger.info(f"Exported {len(experiment_ids)} experiments to {output_path}")

    def export_to_json(
        self, experiment_id: str, output_path: str, include_raw_output: bool = False
    ) -> None:
        """Export experiment results to JSON format.

        Args:
            experiment_id: Experiment identifier
            output_path: Path for output JSON file
            include_raw_output: Whether to include raw model outputs
        """
        result = {
            "experiment_metadata": {},
            "summary_statistics": {},
            "runs": [],
            "analysis": {},
        }

        with self.db_manager.get_connection() as conn:
            # Get experiment metadata
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM experiments WHERE id = ?", (experiment_id,))
            exp_data = cursor.fetchone()

            if exp_data:
                result["experiment_metadata"] = {
                    "id": exp_data["id"],
                    "name": exp_data["name"],
                    "description": exp_data["description"],
                    "created_at": exp_data["created_at"],
                    "status": exp_data["status"],
                    "config": json.loads(exp_data["config_json"]),
                }

            # Get summary statistics
            summary = self.get_experiment_summary(experiment_id)
            if summary:
                result["summary_statistics"] = asdict(summary)
                result["summary_statistics"][
                    "created_at"
                ] = summary.created_at.isoformat()
                result["summary_statistics"][
                    "last_updated"
                ] = summary.last_updated.isoformat()

            # Get all runs
            cursor.execute(
                """
                SELECT
                    r.*,
                    pm.parsing_attempts,
                    pm.fallback_used,
                    pm.confidence_score as parsing_confidence_score
                FROM runs r
                LEFT JOIN parsing_metrics pm ON r.id = pm.run_id
                WHERE r.experiment_id = ?
                ORDER BY r.approach_name, r.sample_index
            """,
                (experiment_id,),
            )

            for row in cursor.fetchall():
                run_data = dict(row)

                # Parse metadata JSON
                if run_data.get("metadata_json"):
                    run_data["metadata"] = json.loads(run_data["metadata_json"])
                    del run_data["metadata_json"]

                # Optionally exclude raw output
                if not include_raw_output and "raw_output" in run_data:
                    run_data["raw_output"] = f"[{len(run_data['raw_output'])} chars]"

                result["runs"].append(run_data)

            # Add analysis
            result["analysis"] = {
                "accuracy_report": self.generate_accuracy_report(experiment_id),
                "approach_comparison": [
                    asdict(comp) for comp in self.compare_approaches(experiment_id)
                ],
                "failure_patterns": self.identify_failure_patterns(
                    experiment_id, limit=10
                ),
            }

        # Save to JSON
        with open(output_path, "w") as f:
            json.dump(result, f, indent=2, default=str)

        logger.info(f"Exported experiment {experiment_id} to {output_path}")

    # ==================== Utility Methods ====================

    def get_experiments_list(
        self, status: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Get list of all experiments.

        Args:
            status: Optional status filter ('running', 'completed', 'failed')

        Returns:
            List of experiment metadata
        """
        with self.db_manager.get_connection() as conn:
            query = "SELECT * FROM experiments"
            params = []

            if status:
                query += " WHERE status = ?"
                params.append(status)

            query += " ORDER BY created_at DESC"

            cursor = conn.cursor()
            cursor.execute(query, params)

            experiments = []
            for row in cursor.fetchall():
                exp = dict(row)
                exp["config"] = json.loads(exp["config_json"])
                del exp["config_json"]
                experiments.append(exp)

            return experiments

    def get_run_details(self, run_id: str) -> Optional[Dict[str, Any]]:
        """Get detailed information for a specific run.

        Args:
            run_id: Run identifier

        Returns:
            Run details or None if not found
        """
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    r.*,
                    pm.parsing_attempts,
                    pm.fallback_used,
                    pm.confidence_score as parsing_confidence_score,
                    pm.extraction_time_ms,
                    pm.error_details
                FROM runs r
                LEFT JOIN parsing_metrics pm ON r.id = pm.run_id
                WHERE r.id = ?
            """,
                (run_id,),
            )

            row = cursor.fetchone()
            if row:
                run_data = dict(row)
                if run_data.get("metadata_json"):
                    run_data["metadata"] = json.loads(run_data["metadata_json"])
                    del run_data["metadata_json"]
                return run_data

            return None

    def cleanup_old_experiments(self, days: int = 30) -> int:
        """Remove experiments older than specified days.

        Args:
            days: Number of days to keep

        Returns:
            Number of experiments deleted
        """
        cutoff_date = datetime.now().timestamp() - (days * 24 * 60 * 60)

        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()

            # Get experiments to delete
            cursor.execute(
                """
                SELECT id FROM experiments
                WHERE julianday('now') - julianday(created_at) > ?
            """,
                (days,),
            )

            exp_ids = [row[0] for row in cursor.fetchall()]

            if exp_ids:
                # Delete experiments (cascades to runs and parsing_metrics)
                placeholders = ",".join(["?" for _ in exp_ids])
                cursor.execute(
                    f"DELETE FROM experiments WHERE id IN ({placeholders})", exp_ids
                )
                conn.commit()

            logger.info(f"Deleted {len(exp_ids)} experiments older than {days} days")
            return len(exp_ids)
